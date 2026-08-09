#!/usr/bin/env python3
"""生成市政道路改造工程成本测算工作簿（openpyxl，原生公式 + 图表）。

用法：
    python cost_estimate.py 输出.xlsx [工程名称]

结构：
    Sheet1 成本测算：按费用构成（直接费=人工/材料/机械，企管/规费/利润/税金）
        含 数量×单价=合价、小计、占比、综合单价公式；
    Sheet2 费用汇总：类别汇总 + 原生饼图（构成占比）+ 柱状图（直接费构成）；
    Sheet3 编制说明：编制依据、取费标准、数据口径。

生成后建议用 recalc.py 重算，让公式带缓存值（LibreOffice）。
"""

import argparse
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
HEADER_FILL = "4472C4"
TOTAL_FILL = "DCE6F1"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_cost_sheet(ws, qty_area):
    """Sheet1 成本测算：主表 + 公式。"""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "某市政道路改造工程成本测算表"
    c.font = Font(name="微软雅黑", size=16, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "编制依据：GB 50500-2013 清单计价规范；材料价：2026 年第 2 期信息价；增值税率 9%；车行道改造面积 {} m²".format(qty_area)
    c.font = Font(name="微软雅黑", size=9, color="808080")
    c.alignment = Alignment(horizontal="center")

    headers = ["序号", "费用项目", "规格/说明", "单位", "数量", "单价(元)", "合价(元)", "占直接费比"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[4].height = 22

    # (序号, 费用项目, 规格/说明, 单位, 数量, 单价, 是否小计行)
    rows = [
        (1, "人工费", "普工+技工综合", "工日", 8640, 265),
        (2, "细粒式沥青混凝土 AC-13", "厚 5cm（车行道）", "t", 2808, 485),
        (3, "中粒式沥青混凝土 AC-20", "厚 7cm（车行道）", "t", 3856, 455),
        (4, "水泥稳定碎石基层", "5% 水泥 厚 18cm", "m³", 3888, 330),
        (5, "花岗岩路缘石", "12.5×30cm", "m", 4800, 68),
        (6, "人行道透水砖", "60×30×6cm", "m²", 4800, 95),
        (7, "机械费", "摊铺/压实/运输综合台班", "台班", 96, 1860),
    ]
    r = 5
    for no, name, spec, unit, qty, price in rows:
        ws.cell(row=r, column=1, value=no)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=spec)
        ws.cell(row=r, column=4, value=unit)
        ws.cell(row=r, column=5, value=qty)
        ws.cell(row=r, column=6, value=price)
        ws.cell(row=r, column=7, value="=E%d*F%d" % (r, r))
        ws.cell(row=r, column=8, value="=G%d/$G$12" % r)
        for ci in range(1, 9):
            cell = ws.cell(row=r, column=ci)
            cell.border = BORDER
            cell.font = Font(name="微软雅黑", size=10)
            if ci in (1, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center")
            if ci == 7:
                cell.number_format = "#,##0.00"
            if ci == 8:
                cell.number_format = "0.0%"
        r += 1

    # 直接费合计 / 企管 / 规费 / 利润 / 税金 / 含税总价 / 综合单价
    r1 = r
    ws.cell(row=r, column=2, value="直接费合计")
    ws.cell(row=r, column=7, value="=SUM(G5:G%d)" % (r - 1))
    ws.cell(row=r, column=8, value="=G%d/G%d" % (r, r))
    r += 1
    ws.cell(row=r, column=2, value="企业管理费")
    ws.cell(row=r, column=3, value="直接费×10%")
    ws.cell(row=r, column=7, value="=G%d*0.10" % r1)
    ws.cell(row=r, column=8, value="=G%d/$G$17" % r)
    r += 1
    ws.cell(row=r, column=2, value="规费")
    ws.cell(row=r, column=3, value="直接费×2%")
    ws.cell(row=r, column=7, value="=G%d*0.02" % r1)
    ws.cell(row=r, column=8, value="=G%d/$G$17" % r)
    r += 1
    profit_row = r
    ws.cell(row=r, column=2, value="利润")
    ws.cell(row=r, column=3, value="(直接费+企管+规费)×7%")
    ws.cell(row=r, column=7, value="=(G%d+G%d+G%d)*0.07" % (r1, r1 + 1, r1 + 2))
    ws.cell(row=r, column=8, value="=G%d/$G$17" % r)
    r += 1
    tax_row = r
    ws.cell(row=r, column=2, value="税金（9%）")
    ws.cell(row=r, column=3, value="(直接费+企管+规费+利润)×9%")
    ws.cell(row=r, column=7, value="=(G%d+G%d+G%d+G%d)*0.09" % (r1, r1 + 1, r1 + 2, profit_row))
    ws.cell(row=r, column=8, value="=G%d/$G$17" % r)
    r += 1
    total_row = r
    ws.cell(row=r, column=2, value="含税总造价")
    ws.cell(row=r, column=7, value="=SUM(G%d:G%d)" % (r1, r - 1))
    ws.cell(row=r, column=8, value="=G%d/G%d" % (r, r))
    r += 1
    ws.cell(row=r, column=2, value="综合单价（元/m²）")
    ws.cell(row=r, column=3, value="含税总造价÷车行道面积")
    ws.cell(row=r, column=4, value="m²")
    ws.cell(row=r, column=5, value=qty_area)
    ws.cell(row=r, column=7, value="=G%d/E%d" % (total_row, r))

    for rr in range(r1, r + 1):
        for ci in range(1, 9):
            cell = ws.cell(row=rr, column=ci)
            cell.border = BORDER
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            if ci in (1, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center")
            if ci == 7:
                cell.number_format = "#,##0.00"
            if ci == 8:
                cell.number_format = "0.0%"

    widths = [6, 26, 30, 8, 10, 12, 14, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A5"
    return r1, profit_row, tax_row, total_row


def build_summary_sheet(ws, r1, profit_row, tax_row, total_row):
    """Sheet2 费用汇总：类别汇总 + 饼图（构成占比）+ 柱状图（直接费构成）。"""
    ws.merge_cells("A1:B1")
    ws["A1"] = "费用汇总"
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=NAVY)

    ws["A3"] = "费用类别"
    ws["B3"] = "金额(元)"
    ws["C3"] = "占含税总价"
    for ci in (1, 2, 3):
        cell = ws.cell(row=3, column=ci)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER

    items = [
        ("直接费", "='成本测算'!G%d" % r1),
        ("企业管理费", "='成本测算'!G%d" % (r1 + 1)),
        ("规费", "='成本测算'!G%d" % (r1 + 2)),
        ("利润", "='成本测算'!G%d" % profit_row),
        ("税金", "='成本测算'!G%d" % tax_row),
        ("含税总造价", "='成本测算'!G%d" % total_row),
    ]
    r = 4
    for name, formula in items:
        ws.cell(row=r, column=1, value=name).border = BORDER
        c = ws.cell(row=r, column=2, value=formula)
        c.number_format = "#,##0.00"
        c.border = BORDER
        c = ws.cell(row=r, column=3, value="=B%d/$B$9" % r)
        c.number_format = "0.0%"
        c.border = BORDER
        if name == "含税总造价":
            for ci in (1, 2, 3):
                ws.cell(row=r, column=ci).font = Font(name="微软雅黑", size=10, bold=True)
                ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor=TOTAL_FILL)
        r += 1

    # 饼图：费用构成占比（不含含税总价行）
    pie = PieChart()
    pie.title = "费用构成占比"
    pie.style = 10
    data = Reference(ws, min_col=2, min_row=3, max_row=8)
    cats = Reference(ws, min_col=1, min_row=4, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = None
    pie.width, pie.height = 13, 8
    ws.add_chart(pie, "E3")

    # 直接费构成明细（引用成本测算表）
    ws["A12"] = "直接费构成明细"
    ws["A12"].font = Font(name="微软雅黑", size=12, bold=True, color=NAVY)
    ws["A13"] = "费用项目"
    ws["B13"] = "合价(元)"
    for ci in (1, 2):
        cell = ws.cell(row=13, column=ci)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER
    for i in range(7):
        rr = 14 + i
        ws.cell(row=rr, column=1, value="='成本测算'!B%d" % (5 + i)).border = BORDER
        c = ws.cell(row=rr, column=2, value="='成本测算'!G%d" % (5 + i))
        c.number_format = "#,##0.00"
        c.border = BORDER

    bar = BarChart()
    bar.type = "col"
    bar.title = "直接费构成对比"
    bar.style = 10
    data = Reference(ws, min_col=2, min_row=13, max_row=20)
    cats = Reference(ws, min_col=1, min_row=14, max_row=20)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.legend = None
    bar.width, bar.height = 13, 8
    ws.add_chart(bar, "E21")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 13


def build_notes_sheet(ws):
    """Sheet3 编制说明。"""
    ws.merge_cells("A1:F1")
    ws["A1"] = "编制说明"
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=NAVY)
    notes = [
        "一、编制依据",
        "1. 《建设工程工程量清单计价规范》GB 50500-2013；",
        "2. 《市政工程消耗量定额》（ZYA 1-31-2015）及当地现行补充定额；",
        "3. 材料价格：2026 年第 2 期当地工程造价信息价，缺项按市场询价；",
        "4. 人工费：综合工日单价按当地发布的人工成本信息调整。",
        "",
        "二、取费标准",
        "1. 企业管理费按直接费 10% 计取；规费按直接费 2% 计取；",
        "2. 利润按（直接费+企业管理费+规费）×7% 计取；",
        "3. 增值税销项税率 9%，按一般计税方法计取。",
        "",
        "三、数据口径",
        "1. 车行道改造面积 21,600 m²（长 1,200m×宽 18m），未含人行道与管网；",
        "2. 沥青混合料用量按设计厚度与压实系数折算；",
        "3. 本表金额为暂估，结算以实际签认工程量与信息价为准。",
        "",
        "四、复核说明",
        "1. 编制：____    复核：____    日期：____；",
        "2. 数据源更新后，用 gaea 重新导出即可同步本表与图表。",
    ]
    r = 2
    for line in notes:
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = Font(name="微软雅黑", size=10)
        r += 1
    ws.column_dimensions["A"].width = 55


def main():
    ap = argparse.ArgumentParser(description="生成市政道路改造工程成本测算工作簿")
    ap.add_argument("output", help="输出 .xlsx 路径")
    ap.add_argument("--area", type=int, default=21600, help="车行道改造面积（m²）")
    args = ap.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "成本测算"
    r1, profit_row, tax_row, total_row = build_cost_sheet(ws, args.area)
    build_summary_sheet(wb.create_sheet("费用汇总"), r1, profit_row, tax_row, total_row)
    build_notes_sheet(wb.create_sheet("编制说明"))
    wb.save(args.output)
    print("saved: %s" % args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
